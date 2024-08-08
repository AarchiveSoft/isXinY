import os.path
import sys

import pandas as pd
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (QApplication, QWidget, QGridLayout, QFileDialog, QLabel, QComboBox, QLineEdit,
                             QPushButton, QFrame)
from openpyxl.styles import PatternFill


class Gui(QWidget):
    """

    :class: :py:class:`Gui(QWidget)`

    This class represents a graphical user interface for an application. It inherits from the `QWidget` class.

    **Properties:**

    - `title`: The title of the GUI window. Default is 'Ist X in Y?'.
    - `window_height`: The height of the GUI window. Default is 900.
    - `window_width`: The width of the GUI window. Default is 450.
    - `icon_path`: The path to the icon file for the GUI window. Not set by default.
    - `layout`: The layout of the GUI window. Initialized as a `QGridLayout` object.

    **Methods:**

    - `__init__()` - Initializes the `Gui` object and sets the properties.
    - `initUi()` - Initializes the user interface components of the GUI.
    - `select_file(type)` - Opens a file dialog to select a file and sets the corresponding path based on the file type (1 for 'is' path, 2 for 'in' path).
    - `select_out_path()` - Opens a dialog to select a directory to save the output file.
    - `confirm()` - Triggers the execution of a comparison process using the selected files and outputs the result to the selected directory.

    Example Usage:
    ```python
    gui = Gui()
    gui.initUi()
    gui.show()
    ```
    """
    def __init__(self):
        """
        Initializes the object and sets the initial values for the following attributes:

        :param self: The object instance.
        :type self: object

        :return: None
        :rtype: None

        :Example:

        >>> app = MyApp()
        >>> app.__init__()

        """
        super().__init__()
        self.title = 'Ist X in Y?'
        self.window_height = 900
        self.window_width = 450
        self.icon_path = ""

        self.initUi()

    def initUi(self):
        """
        Initialize the user interface.

        :return: None
        """
        self.setWindowTitle(self.title)
        self.layout = QGridLayout()

        self.inTitleLabel = QLabel("<b>INPUT</b>", self)
        self.inTitleLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(self.inTitleLabel, 0, 0, 1, -1)

        self.isColLabel = QLabel("Sind EAN's von Spalte: ", self)
        self.layout.addWidget(self.isColLabel, 1, 0)
        self.isColumnSel = QComboBox()
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            self.isColumnSel.addItem(column)
        self.layout.addWidget(self.isColumnSel, 1, 1)
        self.isPathLabel = QLabel("der Liste: ", self)
        self.layout.addWidget(self.isPathLabel, 2, 0)
        self.isPathInput = QLineEdit("")
        self.layout.addWidget(self.isPathInput, 3, 0)
        self.isPathButton = QPushButton("Durchsuchen...", self)
        self.isPathButton.clicked.connect(lambda: self.select_file(1))
        self.layout.addWidget(self.isPathButton, 3, 1)
        self.inColLabel = QLabel("Vorhanden in Spalte: ", self)
        self.layout.addWidget(self.inColLabel, 4, 0)
        self.inColumnSel = QComboBox()
        for column in ['A', 'B', 'C', 'D', 'E', 'F']:
            self.inColumnSel.addItem(column)
        self.layout.addWidget(self.inColumnSel, 4, 1)
        self.inPathLabel = QLabel("der Liste: ", self)
        self.layout.addWidget(self.inPathLabel, 5, 0)
        self.inPathInput = QLineEdit("")
        self.layout.addWidget(self.inPathInput, 6, 0)
        self.inPathButton = QPushButton("Durchsuchen...", self)
        self.inPathButton.clicked.connect(lambda: self.select_file(2))
        self.layout.addWidget(self.inPathButton, 6, 1)
        self.small_list_path = ""
        self.isListCol = self.isColumnSel.currentIndex()
        self.large_list_path = ""
        self.inListCol = self.inColumnSel.currentIndex()
        self.output_path = ""

        self.h_line1 = QFrame()
        self.h_line1.setFrameShape(QFrame.Shape.HLine)
        self.layout.addWidget(self.h_line1, 7, 0, 1, -1)

        self.outTitleLabel = QLabel("<b>OUTPUT</b>", self)
        self.outTitleLabel.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.layout.addWidget(self.outTitleLabel, 8, 0, 1, -1)

        self.outLabel = QLabel("Output speichern unter: ", self)
        self.layout.addWidget(self.outLabel, 9, 0)
        self.outPathInput = QLineEdit("", self)
        self.layout.addWidget(self.outPathInput, 10, 0)
        self.outButton = QPushButton("Durchsuchen...", self)
        self.outButton.clicked.connect(self.select_out_path)
        self.layout.addWidget(self.outButton, 10, 1)
        self.outNameLabel = QLabel("Output File Name: ", self)
        self.layout.addWidget(self.outNameLabel, 11, 0)
        self.outNameInput = QLineEdit("", self)
        self.layout.addWidget(self.outNameInput, 12, 0, 1, -1)

        self.h_line2 = QFrame()
        self.h_line2.setFrameShape(QFrame.Shape.HLine)
        self.layout.addWidget(self.h_line2, 13, 0, 1, -1)

        self.confirm_button = QPushButton("Bestätigen")
        self.confirm_button.setStyleSheet(
            "font-weight: bold;"
            "padding: 4px"
        )
        self.confirm_button.clicked.connect(self.confirm)
        self.layout.addWidget(self.confirm_button, 14, 0, 1, -1)

        self.setLayout(self.layout)

    def select_file(self, type):
        """
        :param type: The type of file path to select. Should be either 1 or 2.
        :return: None

        Selects a file using a file dialog and sets the corresponding file paths based on the provided type parameter.

        Example usage:
            select_file(1)  # Select file for 'is' path
            select_file(2)  # Select file for 'in' path
        """
        file_dialog = QFileDialog()
        file_path, _ = QFileDialog.getOpenFileName(file_dialog, "Excel Datei wählen <b>(muss '.xlsx' sein, wenn nötig "
                                                                "erst konvertieren)</b>",
                                                   "", "Excel Files ("
                                                       "*.xlsx)")
        if file_path:
            original_file_name = os.path.basename(file_path)
            split_file_name = original_file_name.rsplit(".", 1)
            file_name = f'{split_file_name[0]}_status.{split_file_name[1]}'
            if type == 1:  # for 'is' path
                self.small_list_path = file_path
                self.isPathInput.setText(self.small_list_path)
                self.outNameInput.setText(file_name)
            elif type == 2:  # for 'in' path
                self.large_list_path = file_path
                self.inPathInput.setText(self.large_list_path)
            else:
                print("Something went wrong")

    def select_out_path(self):
        """
        Selects the output path by opening a directory dialog and setting the selected directory as the output path.

        :return: None
        """
        directory = QFileDialog.getExistingDirectory(self, "Speicherort wählen")
        if directory:
            self.output_path = directory
        return

    def confirm(self):
        """
        Confirm method executes the Compare.execute method with the specified parameters. It compares two lists based on the provided paths and column indexes, and writes the result to the output path.

        :return: None
        """
        Compare.execute(self.small_list_path, self.isListCol, self.large_list_path, self.inListCol, self. output_path)
        return


class Compare:
    """
    Class to compare values between two Excel files.

    Attributes:
        None

    Methods:
        execute(small_list_path, small_list_col, large_list_path, large_list_col, output_path):
            Executes the comparison between two Excel files and saves the result to a new Excel file.
    """

    def execute(self, small_list_path, small_list_col, large_list_path, large_list_col, output_path):
        """
        :param small_list_path: The path of the small list Excel file.
        :param small_list_col: The column index of the value to be checked in the small list.
        :param large_list_path: The path of the large list Excel file.
        :param large_list_col: The column index of the value to be checked in the large list.
        :param output_path: The path to save the modified Excel file.
        :return: None

        This method executes a comparison between a small list and a large list Excel file. It reads the two files, selects the specified columns, and checks if the values in the small list exist in the large list. It updates the small list with a new column indicating the result of the check and saves the modified file with conditional formatting. Finally, it prints a completion message.

        Example usage:

        execute("small_list.xlsx", 2, "large_list.xlsx", 3, "output.xlsx")
        """
        # Function to read an Excel file and handle exceptions
        def read_excel_file(path):
            """
            Read an Excel file from the given path.

            :param path: The path of the Excel file to be read.
            :return: The pandas DataFrame containing the data read from the Excel file.
            :raises: Exception if there is an error reading the file.
            """
            try:
                df = pd.read_excel(path, engine='openpyxl')
                return df
            except Exception as e:
                print(f"Error reading {path}: {e}")
                raise

        # Read the Excel files
        small_df = read_excel_file(small_list_path)
        large_df = read_excel_file(large_list_path)

        # Print the first few rows to ensure data is read correctly
        print("Small list preview:")
        print(small_df.head())
        print("\nLarge list preview:")
        print(large_df.head())

        # Assume the value to check is in a specific column
        # Adjust the column index if necessary
        small_values = small_df.iloc[:, small_list_col]  # select column (EAN)
        large_values = large_df.iloc[:, large_list_col]  # select column (EAN)

        # Print the values to be compared
        print("\nValues from small list (first 5):")
        print(small_values.head())
        print("\nValues from large list (first 5):")
        print(large_values.head())

        # Create a new column 'status' to store the result of the check
        def check_value_in_large_list(value, large_values):
            """
            Check if a value exists in a large list.

            :param value: The value to check.
            :param large_values: The large list to search in.
            :type value: Any
            :type large_values: list
            :return: 'found' if the value is found, 'not found' otherwise.
            :rtype: str
            """
            if value in large_values.values:
                print(f"Comparing: {value} - Found in large list")
                return 'found'
            else:
                print(f"Comparing: {value} - Not found in large list")
                return 'not found'

        small_df['status'] = small_values.apply(lambda x: check_value_in_large_list(x, large_values))

        # Save the modified DataFrame back to an Excel file with conditional formatting
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        small_df.to_excel(writer, index=False)

        # Load the workbook and select the active sheet
        workbook = writer.book
        sheet = workbook.active

        # Define fill colors
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Apply the fill colors based on the status
        for row in range(2, len(small_df) + 2):  # Start from the second row because the first row is the header
            status = sheet[f'E{row}'].value  # Assuming the 'status' column is the fifth column
            fill = green_fill if status == 'found' else red_fill
            for col in range(1, 6):  # Adjust the range if there are more columns to color
                sheet.cell(row=row, column=col).fill = fill

        # Save the workbook
        workbook.save(output_path)
        workbook.close()

        print("Script completed successfully.")


if __name__ == "__main__":
    app = QApplication(sys.argv)

    window = Gui()  # Your QWidget subclass
    window.show()

    sys.exit(app.exec())
